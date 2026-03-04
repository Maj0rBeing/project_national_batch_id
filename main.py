from __future__ import annotations

import argparse
import csv
import json
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

from PIL import Image, ImageDraw, ImageFont, ImageOps

# Module-level font cache: (font_path_or_None, size) -> font object
_font_cache: Dict[tuple, ImageFont.ImageFont] = {}


# ==========================
# CONFIG
# ==========================
CSV_FILE = "id_data.csv"
TEMPLATE_PATH = "id_template.png"
PHOTO_FOLDER = "photos"
OUTPUT_FOLDER = "output"
REPORT_FOLDER = "report"
DISTRICT_REGISTRY_FOLDER = "district_id_registry"
DISTRICT_MIN = 1
DISTRICT_MAX = 15

# Photo box (top-left corner where the photo will be pasted)
PHOTO_SIZE = (200, 200)
CONTENT_CENTER_X = 400
PHOTO_TOP_Y = 315
PHOTO_POSITION = (CONTENT_CENTER_X - (PHOTO_SIZE[0] // 2), PHOTO_TOP_Y)

# Text block starts below the photo and flows downward.
TEXT_START_X = CONTENT_CENTER_X - 130
TEXT_START_Y = PHOTO_POSITION[1] + PHOTO_SIZE[1] + 16

# Wrap area width for all text blocks
WRAP_MAX_WIDTH = 350
SECTION_GAP = 25

# Colors (hex supported by Pillow)
COLOR_BLACK = "#ffffff"
COLOR_RED   = "#FF0000"

# Fonts
# Option A (recommended): Put your font files inside ./fonts and set paths here.
# Example:
# FONT_NAME_PATH = os.path.join("fonts", "Brigends Expanded.otf")
# If you don't have custom fonts, keep None and it will fall back to arial.
FONT_NAME_PATH: Optional[str] = None
FONT_ID_PATH: Optional[str] = None
FONT_ROLE_PATH: Optional[str] = None
FONT_SMALL_PATH: Optional[str] = None

FONT_NAME_SIZE = 60
FONT_ID_SIZE = 38
FONT_ROLE_SIZE = 55
FONT_SMALL_SIZE = 25


# ==========================
# HELPERS
# ==========================
def safe_filename(text: str) -> str:
    text = text.strip().replace(" ", "_")
    text = re.sub(r"[^A-Za-z0-9_\\-]+", "", text)
    return text or "unknown"


def normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def normalize_person_key(firstname: str, lastname: str) -> str:
    return f"{normalize_spaces(firstname).lower()}|{normalize_spaces(lastname).lower()}"


def extract_district_number(district_text: str) -> int:
    match = re.search(r"(\d+)", normalize_spaces(district_text))
    if not match:
        raise ValueError(f"District value '{district_text}' does not contain a number")

    district_number = int(match.group(1))
    if not (DISTRICT_MIN <= district_number <= DISTRICT_MAX):
        raise ValueError(f"District number {district_number} is out of range ({DISTRICT_MIN}-{DISTRICT_MAX})")
    return district_number


def district_registry_path(registry_folder: str, district_number: int) -> str:
    return os.path.join(registry_folder, f"district_{district_number:02d}.json")


def default_district_registry(district_number: int) -> Dict[str, Any]:
    return {
        "district_number": district_number,
        "district_label": f"District {district_number}",
        "last_id_number": 0,
        "records": {},
    }


def save_district_registry(registry_folder: str, district_number: int, data: Dict[str, Any]) -> None:
    os.makedirs(registry_folder, exist_ok=True)
    path = district_registry_path(registry_folder, district_number)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=True)


def load_district_registry(registry_folder: str, district_number: int) -> Dict[str, Any]:
    path = district_registry_path(registry_folder, district_number)
    if not os.path.exists(path):
        data = default_district_registry(district_number)
        save_district_registry(registry_folder, district_number, data)
        return data

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    data.setdefault("district_number", district_number)
    data.setdefault("district_label", f"District {district_number}")
    data.setdefault("last_id_number", 0)
    data.setdefault("records", {})
    if not isinstance(data["records"], dict):
        data["records"] = {}
    return data


def initialize_district_registries(registry_folder: str) -> None:
    os.makedirs(registry_folder, exist_ok=True)
    for district_number in range(DISTRICT_MIN, DISTRICT_MAX + 1):
        path = district_registry_path(registry_folder, district_number)
        if not os.path.exists(path):
            save_district_registry(registry_folder, district_number, default_district_registry(district_number))


def format_assigned_id(district_number: int, sequence_number: int) -> str:
    return f"D{district_number:02d}-{sequence_number:03d}"


def district_output_folder(root_output_folder: str, district_number: int, batch_timestamp: str) -> str:
    return os.path.join(
        root_output_folder,
        f"district_{district_number:02d}",
        f"batch_{batch_timestamp}",
    )


def district_report_folder(root_report_folder: str, district_number: int) -> str:
    return os.path.join(root_report_folder, f"district_{district_number:02d}")


def district_report_file_path(root_report_folder: str, district_number: int) -> str:
    return os.path.join(
        district_report_folder(root_report_folder, district_number),
        f"district_{district_number:02d}_report.xlsx",
    )


def create_batch_report_buckets() -> Dict[int, List[Dict[str, str]]]:
    return {district_number: [] for district_number in range(DISTRICT_MIN, DISTRICT_MAX + 1)}


def write_excel_reports(
    root_report_folder: str,
    batch_timestamp: str,
    rows_by_district: Dict[int, List[Dict[str, str]]],
) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import PatternFill
    except ModuleNotFoundError as e:
        raise ModuleNotFoundError(
            "Missing dependency 'openpyxl'. Install it (for example: pip install -r requirements.txt) "
            "to generate district Excel reports."
        ) from e

    headers = [
        "batch_timestamp",
        "district_number",
        "district_text",
        "assigned_id",
        "firstname",
        "lastname",
        "role",
        "school",
        "photo",
        "output_file",
        "rendered_at",
    ]
    # Requested uniqueness key for report rows (across batches).
    unique_key_headers = ["assigned_id", "firstname", "lastname"]
    NEW_ENTRY_FILL_COLOR = "92D050"  # bright green — reserved for brand-new rows only
    batch_fill_palette = [
        "FFF2CC",  # light yellow
        "DDEBF7",  # light blue
        "E2F0D9",  # light green
        "FCE4D6",  # light orange
        "EAD1DC",  # light pink
        "D9EAD3",  # pale green
        "D0E0E3",  # pale cyan
        "F4CCCC",  # pale red
    ]

    for district_number in range(DISTRICT_MIN, DISTRICT_MAX + 1):
        district_rows = rows_by_district.get(district_number, [])
        if not district_rows:
            continue

        district_folder = district_report_folder(root_report_folder, district_number)
        os.makedirs(district_folder, exist_ok=True)
        report_path = district_report_file_path(root_report_folder, district_number)

        if os.path.exists(report_path):
            workbook = load_workbook(report_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"District {district_number}"

        # Ensure header row exists and matches expected columns.
        existing_headers = [sheet.cell(row=1, column=i + 1).value for i in range(len(headers))]
        if existing_headers != headers:
            for i, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=i, value=header)

        header_to_col = {header: index + 1 for index, header in enumerate(headers)}

        def norm(value: Any) -> str:
            return str(value or "").strip().lower()

        existing_index: Dict[tuple, int] = {}
        existing_by_assigned_id: Dict[str, int] = {}
        rendered_at_col = header_to_col["rendered_at"]
        output_file_col = header_to_col["output_file"]
        for row_idx in range(2, sheet.max_row + 1):
            key = tuple(
                norm(sheet.cell(row=row_idx, column=header_to_col[h]).value)
                for h in unique_key_headers
            )
            if any(key):
                existing_index[key] = row_idx
            assigned_id_key = norm(sheet.cell(row=row_idx, column=header_to_col["assigned_id"]).value)
            if assigned_id_key:
                existing_by_assigned_id[assigned_id_key] = row_idx

        # Back-populate rendered_at for historical rows that predate this feature.
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=rendered_at_col).value:
                continue
            output_file = str(sheet.cell(row=row_idx, column=output_file_col).value or "").strip()
            if output_file and os.path.exists(output_file):
                mtime = datetime.fromtimestamp(os.path.getmtime(output_file)).strftime("%Y%m%d_%H%M%S")
                sheet.cell(row=row_idx, column=rendered_at_col, value=mtime)

        new_row_indices: Set[int] = set()
        for row in district_rows:
            row_key = tuple(norm(row.get(h, "")) for h in unique_key_headers)
            target_row_idx = existing_index.get(row_key)

            if target_row_idx is None:
                # Better fallback: assigned_id is already district-unique and persistent.
                # If names were corrected/renamed, update the existing assigned_id row instead of duplicating it.
                assigned_id_only = norm(row.get("assigned_id", ""))
                target_row_idx = existing_by_assigned_id.get(assigned_id_only)

            if target_row_idx is None:
                target_row_idx = sheet.max_row + 1
                new_row_indices.add(target_row_idx)

            existing_index[row_key] = target_row_idx
            assigned_id_only = norm(row.get("assigned_id", ""))
            if assigned_id_only:
                existing_by_assigned_id[assigned_id_only] = target_row_idx

            for header in headers:
                if header == "rendered_at":
                    # Preserve once set — never overwrite the original render timestamp.
                    cell = sheet.cell(row=target_row_idx, column=header_to_col[header])
                    if not cell.value:
                        cell.value = row.get(header, "")
                else:
                    sheet.cell(row=target_row_idx, column=header_to_col[header], value=row.get(header, ""))

        # Color rows:
        #   - Brand-new rows (added to report for the first time) → bright green
        #   - All other rows → grouped by rendered_at (when the PNG was originally created)
        new_entry_fill = PatternFill(fill_type="solid", start_color=NEW_ENTRY_FILL_COLOR, end_color=NEW_ENTRY_FILL_COLOR)
        batch_fill_map: Dict[str, Any] = {}
        next_fill_index = 0
        no_fill = PatternFill(fill_type=None)
        for row_idx in range(2, sheet.max_row + 1):
            if row_idx in new_row_indices:
                fill = new_entry_fill
            else:
                rendered_value = str(sheet.cell(row=row_idx, column=rendered_at_col).value or "").strip()
                if not rendered_value:
                    fill = no_fill
                else:
                    if rendered_value not in batch_fill_map:
                        color = batch_fill_palette[next_fill_index % len(batch_fill_palette)]
                        batch_fill_map[rendered_value] = PatternFill(
                            fill_type="solid",
                            start_color=color,
                            end_color=color,
                        )
                        next_fill_index += 1
                    fill = batch_fill_map[rendered_value]

            for col_idx in range(1, len(headers) + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = fill

        try:
            workbook.save(report_path)
            print(f"Report: {report_path}")
        except PermissionError:
            print(f"[SKIP] {report_path} — file is open in another program, close it and re-run.")


def get_or_create_district_id(
    registry_folder: str,
    district_number: int,
    firstname: str,
    lastname: str,
    photo_filename: str,
) -> str:
    registry = load_district_registry(registry_folder, district_number)
    records = registry.get("records", {})
    person_key = normalize_person_key(firstname, lastname)
    existing = records.get(person_key)

    if isinstance(existing, dict) and existing.get("id_number"):
        id_number = int(existing["id_number"])
        existing["id"] = format_assigned_id(district_number, id_number)
        existing["firstname"] = firstname
        existing["lastname"] = lastname
        existing["photo"] = photo_filename
        records[person_key] = existing
        registry["records"] = records
        registry["last_id_number"] = max(int(registry.get("last_id_number", 0)), id_number)
        save_district_registry(registry_folder, district_number, registry)
        return str(existing["id"])

    max_existing = int(registry.get("last_id_number", 0))
    for record in records.values():
        if isinstance(record, dict):
            try:
                max_existing = max(max_existing, int(record.get("id_number", 0)))
            except Exception:
                pass

    next_id_number = max_existing + 1
    assigned_id = format_assigned_id(district_number, next_id_number)
    records[person_key] = {
        "firstname": firstname,
        "lastname": lastname,
        "photo": photo_filename,
        "id_number": next_id_number,
        "id": assigned_id,
    }
    registry["records"] = records
    registry["last_id_number"] = next_id_number
    save_district_registry(registry_folder, district_number, registry)
    return assigned_id


def assign_id_in_memory(
    registry: Dict[str, Any],
    district_number: int,
    firstname: str,
    lastname: str,
    photo_filename: str,
) -> str:
    """Assign or retrieve a district ID using an in-memory registry dict (no file I/O)."""
    records = registry.setdefault("records", {})
    person_key = normalize_person_key(firstname, lastname)
    existing = records.get(person_key)

    if isinstance(existing, dict) and existing.get("id_number"):
        id_number = int(existing["id_number"])
        existing["id"] = format_assigned_id(district_number, id_number)
        existing["firstname"] = firstname
        existing["lastname"] = lastname
        existing["photo"] = photo_filename
        registry["last_id_number"] = max(int(registry.get("last_id_number", 0)), id_number)
        return str(existing["id"])

    max_existing = int(registry.get("last_id_number", 0))
    for record in records.values():
        if isinstance(record, dict):
            try:
                max_existing = max(max_existing, int(record.get("id_number", 0)))
            except Exception:
                pass

    next_id_number = max_existing + 1
    assigned_id = format_assigned_id(district_number, next_id_number)
    records[person_key] = {
        "firstname": firstname,
        "lastname": lastname,
        "photo": photo_filename,
        "id_number": next_id_number,
        "id": assigned_id,
    }
    registry["last_id_number"] = next_id_number
    return assigned_id


def get_csv_value(row: Dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value:
            return str(value).strip()
    return ""


def is_athlete_role(role: str) -> bool:
    return normalize_spaces(role).lower() == "athlete"


def load_font(font_path: Optional[str], size: int) -> ImageFont.ImageFont:
    key = (font_path, size)
    if key in _font_cache:
        return _font_cache[key]

    if font_path and os.path.exists(font_path):
        font = ImageFont.truetype(font_path, size)
    else:
        try:
            font = ImageFont.truetype("arial.ttf", size)
        except Exception:
            font = ImageFont.load_default()

    _font_cache[key] = font
    return font


def text_width(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> int:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0]


def text_height(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> int:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[3] - bbox[1]


def draw_centered_text(
    draw: ImageDraw.ImageDraw,
    center_x: int,
    y: int,
    text: str,
    font: ImageFont.ImageFont,
    fill: str,
) -> None:
    x = center_x - (text_width(draw, text, font) // 2)
    draw.text((x, y), text, fill=fill, font=font)


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> List[str]:
    # Word-wrap by measuring width
    words = text.split()
    if not words:
        return [""]

    lines: List[str] = []
    line: List[str] = []

    for w in words:
        test = " ".join(line + [w])
        if text_width(draw, test, font) <= max_width:
            line.append(w)
        else:
            if line:
                lines.append(" ".join(line))
                line = [w]
            else:
                # Single word longer than max_width: hard-cut
                lines.append(w)
                line = []

    if line:
        lines.append(" ".join(line))

    return lines


def fit_font_size(
    draw: ImageDraw.ImageDraw,
    text: str,
    font_path: Optional[str],
    start_size: int,
    max_width: int,
    min_size: int = 14,
) -> ImageFont.ImageFont:
    size = start_size
    while size >= min_size:
        font = load_font(font_path, size)
        if text_width(draw, text, font) <= max_width:
            return font
        size -= 1
    return load_font(font_path, min_size)


def open_photo_correct_orientation(path: str) -> Image.Image:
    # Fix EXIF rotation AND convert to RGBA to avoid paste issues
    img = Image.open(path)
    img = ImageOps.exif_transpose(img)
    return img.convert("RGBA")


# ==========================
# CORE
# ==========================
def create_id_card(
    firstname: str,
    lastname: str,
    role: str,
    school: str,
    district: str,
    assigned_id: str,
    photo_filename: str,
    template_image: Image.Image,
    photo_folder: str,
    output_folder: str,
) -> str:
    full_name = f"{firstname} {lastname}".strip()

    template = template_image.copy()
    draw = ImageDraw.Draw(template)

    font_name = fit_font_size(
        draw=draw,
        text=full_name.upper(),
        font_path=FONT_NAME_PATH,
        start_size=FONT_NAME_SIZE,
        max_width=WRAP_MAX_WIDTH,
        min_size=28,
    )
    font_id = load_font(FONT_ID_PATH, FONT_ID_SIZE)
    font_role = fit_font_size(
        draw=draw,
        text=str(role).upper(),
        font_path=FONT_ROLE_PATH,
        start_size=FONT_ROLE_SIZE,
        max_width=WRAP_MAX_WIDTH,
        min_size=20,
    )
    font_small = load_font(FONT_SMALL_PATH, FONT_SMALL_SIZE)

    y = TEXT_START_Y

    name_lines = wrap_text(draw, full_name.upper(), font_name, max_width=WRAP_MAX_WIDTH)
    for ln in name_lines[:2]:
        draw_centered_text(draw, CONTENT_CENTER_X, y, ln, font_name, COLOR_BLACK)
        y += text_height(draw, ln, font_name) + 2
    y += SECTION_GAP

    if is_athlete_role(role):
        id_line = f"ID: {assigned_id}"
        draw_centered_text(draw, CONTENT_CENTER_X, y, id_line, font_id, COLOR_BLACK)
        y += text_height(draw, id_line, font_id) + SECTION_GAP

    role_line = str(role).upper()
    draw_centered_text(draw, CONTENT_CENTER_X, y, role_line, font_role, COLOR_RED)
    y += text_height(draw, role_line, font_role) + SECTION_GAP

    school_text = f"SCHOOL: {school}".upper().strip()
    school_lines = wrap_text(draw, school_text, font_small, WRAP_MAX_WIDTH)
    for ln in school_lines[:3]:
        draw_centered_text(draw, CONTENT_CENTER_X, y, ln, font_small, COLOR_BLACK)
        y += text_height(draw, ln, font_small) + 2
    y += 4

    try:
        district_display = str(extract_district_number(district))
    except Exception:
        district_display = str(district).strip()
    district_text = f"DISTRICT: {district_display}".upper().strip()
    district_lines = wrap_text(draw, district_text, font_small, WRAP_MAX_WIDTH)
    for ln in district_lines[:3]:
        draw_centered_text(draw, CONTENT_CENTER_X, y, ln, font_small, COLOR_BLACK)
        y += text_height(draw, ln, font_small) + 2

    # Photo
    photo_path = os.path.join(photo_folder, photo_filename) if photo_filename else ""
    if photo_path and os.path.exists(photo_path):
        try:
            photo = open_photo_correct_orientation(photo_path)

            # If still rotated weird (no EXIF / wrong EXIF), optional heuristic:
            # If it's very tall vs wide, keep; if it's sideways and looks wrong, you can rotate:
            # if photo.width > photo.height:
            #     photo = photo.rotate(90, expand=True)

            photo = photo.resize(PHOTO_SIZE)
            template.paste(photo, PHOTO_POSITION, mask=photo)
        except Exception as e:
            print(f"[WARN] Photo error for '{photo_path}': {e}")
    else:
        if photo_filename:
            print(f"[WARN] Photo not found: {photo_path}")

    # Save output
    os.makedirs(output_folder, exist_ok=True)
    out_name = safe_filename(f"{firstname}_{lastname}").upper() + ".png"
    out_path = os.path.join(output_folder, out_name)
    template.convert("RGB").save(out_path)
    return out_path


def validate_photos(rows: List[Dict[str, str]], photo_folder: str) -> None:
    missing: List[str] = []
    seen: Set[str] = set()
    for row in rows:
        photo = get_csv_value(row, "Photo", "photo", "PHOTO")
        if photo and photo not in seen:
            seen.add(photo)
            if not os.path.exists(os.path.join(photo_folder, photo)):
                missing.append(photo)
    if missing:
        print(f"[WARN] {len(missing)} photo(s) not found in '{photo_folder}':")
        for name in missing:
            print(f"       - {name}")


def batch_generate_id_cards(csv_file: str, district_filter: Optional[int] = None) -> None:
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Missing template: {TEMPLATE_PATH} (put id_template.png in the project root)")
    initialize_district_registries(DISTRICT_REGISTRY_FOLDER)
    batch_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_rows_by_district = create_batch_report_buckets()
    print(f"Batch timestamp: {batch_timestamp}")

    # IMPORTANT: utf-8-sig removes BOM (\ufeff) from 'firstname'
    with open(csv_file, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        print(f"Detected headers: {reader.fieldnames}")
        all_rows = list(reader)

    total = len(all_rows)
    print(f"Total rows: {total}")

    # Photo pre-validation
    validate_photos(all_rows, PHOTO_FOLDER)

    # Load all district registries into memory once
    registries: Dict[int, Dict[str, Any]] = {
        n: load_district_registry(DISTRICT_REGISTRY_FOLDER, n)
        for n in range(DISTRICT_MIN, DISTRICT_MAX + 1)
    }
    modified_districts: Set[int] = set()

    # Detect districts with no report file — rows for these districts bypass processed=YES
    missing_report_districts: Set[int] = {
        n for n in range(DISTRICT_MIN, DISTRICT_MAX + 1)
        if not os.path.exists(district_report_file_path(REPORT_FOLDER, n))
    }
    if missing_report_districts:
        print(f"[INFO] Missing report file(s) for district(s): {sorted(missing_report_districts)} — will rebuild from CSV.")

    # Open template once
    template_image = Image.open(TEMPLATE_PATH).convert("RGBA")

    stats = {"generated": 0, "skipped": 0, "rebuilt": 0, "card_exists": 0, "warnings": 0, "errors": 0}

    for i, row in enumerate(all_rows, start=1):
        try:
            firstname = get_csv_value(row, "firstname", "Firstname", "FIRSTNAME")
            lastname = get_csv_value(row, "lastname", "Lastname", "LASTNAME")
            role = get_csv_value(row, "Role", "role", "ROLE")
            photo = get_csv_value(row, "Photo", "photo", "PHOTO")
            district = get_csv_value(row, "District", "district", "distrct", "Distrct", "DISTRICT")
            school = get_csv_value(row, "School", "school", "SCHOOL")

            if not firstname and not lastname:
                print(f"[WARN] Row {i}/{total}: missing firstname/lastname - skipping")
                stats["warnings"] += 1
                continue

            if not district:
                print(f"[WARN] Row {i}/{total}: missing district - skipping")
                stats["warnings"] += 1
                continue

            try:
                district_number = extract_district_number(district)
            except ValueError as e:
                print(f"[WARN] Row {i}/{total}: {e} - skipping")
                stats["warnings"] += 1
                continue

            if district_filter is not None and district_number != district_filter:
                continue

            processed_flag = get_csv_value(row, "processed", "Processed", "PROCESSED")
            if processed_flag.upper() == "YES":
                if district_number not in missing_report_districts:
                    stats["skipped"] += 1
                    continue
                stats["rebuilt"] += 1

            assigned_id = ""
            if is_athlete_role(role):
                assigned_id = assign_id_in_memory(
                    registry=registries[district_number],
                    district_number=district_number,
                    firstname=firstname,
                    lastname=lastname,
                    photo_filename=photo,
                )
                modified_districts.add(district_number)

            out_folder = district_output_folder(OUTPUT_FOLDER, district_number, batch_timestamp)
            expected_path = os.path.join(out_folder, safe_filename(f"{firstname}_{lastname}").upper() + ".png")

            if os.path.exists(expected_path):
                print(f"[{i}/{total}] Exists (skipped): {expected_path}")
                stats["card_exists"] += 1
                output_path = expected_path
                rendered_at = datetime.fromtimestamp(os.path.getmtime(output_path)).strftime("%Y%m%d_%H%M%S")
            else:
                output_path = create_id_card(
                    firstname=firstname,
                    lastname=lastname,
                    role=role,
                    school=school,
                    district=district,
                    assigned_id=assigned_id,
                    photo_filename=photo,
                    template_image=template_image,
                    photo_folder=PHOTO_FOLDER,
                    output_folder=out_folder,
                )
                print(f"[{i}/{total}] Saved: {output_path}")
                stats["generated"] += 1
                rendered_at = batch_timestamp
            report_rows_by_district[district_number].append(
                {
                    "batch_timestamp": batch_timestamp,
                    "district_number": str(district_number),
                    "district_text": district,
                    "assigned_id": assigned_id,
                    "firstname": firstname,
                    "lastname": lastname,
                    "role": role,
                    "school": school,
                    "photo": photo,
                    "output_file": output_path,
                    "rendered_at": rendered_at,
                }
            )

        except Exception as e:
            print(f"[ERROR] Row {i}/{total} failed: {e}")
            stats["errors"] += 1

    # Save only modified registries
    for district_number in modified_districts:
        save_district_registry(DISTRICT_REGISTRY_FOLDER, district_number, registries[district_number])

    write_excel_reports(REPORT_FOLDER, batch_timestamp, report_rows_by_district)

    print(
        f"\n=== Done === generated={stats['generated']}  exists(skipped)={stats['card_exists']}"
        f"  skipped(processed)={stats['skipped']}  rebuilt(missing report)={stats['rebuilt']}"
        f"  warnings={stats['warnings']}  errors={stats['errors']}"
    )


def format_existing_reports(root_report_folder: str) -> None:
    """Migrate all existing district report files: add rendered_at column and reapply color coding."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ModuleNotFoundError as e:
        raise ModuleNotFoundError(
            "Missing dependency 'openpyxl'. Install it (for example: pip install -r requirements.txt)."
        ) from e

    headers = [
        "batch_timestamp",
        "district_number",
        "district_text",
        "assigned_id",
        "firstname",
        "lastname",
        "role",
        "school",
        "photo",
        "output_file",
        "rendered_at",
    ]
    NEW_ENTRY_FILL_COLOR = "92D050"
    batch_fill_palette = [
        "FFF2CC",  # light yellow
        "DDEBF7",  # light blue
        "E2F0D9",  # light green
        "FCE4D6",  # light orange
        "EAD1DC",  # light pink
        "D9EAD3",  # pale green
        "D0E0E3",  # pale cyan
        "F4CCCC",  # pale red
    ]

    found = 0
    for district_number in range(DISTRICT_MIN, DISTRICT_MAX + 1):
        report_path = district_report_file_path(root_report_folder, district_number)
        if not os.path.exists(report_path):
            continue

        found += 1
        workbook = load_workbook(report_path)
        sheet = workbook.active

        # Ensure all headers are present (adds rendered_at column if missing).
        existing_headers = [sheet.cell(row=1, column=i + 1).value for i in range(len(headers))]
        if existing_headers != headers:
            for i, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=i, value=header)

        header_to_col = {header: index + 1 for index, header in enumerate(headers)}
        rendered_at_col = header_to_col["rendered_at"]
        output_file_col = header_to_col["output_file"]

        # Back-populate rendered_at from file mtime for any row that doesn't have it yet.
        populated = 0
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=rendered_at_col).value:
                continue
            output_file = str(sheet.cell(row=row_idx, column=output_file_col).value or "").strip()
            if output_file and os.path.exists(output_file):
                mtime = datetime.fromtimestamp(os.path.getmtime(output_file)).strftime("%Y%m%d_%H%M%S")
                sheet.cell(row=row_idx, column=rendered_at_col, value=mtime)
                populated += 1

        # Reapply color coding grouped by rendered_at across the full sheet.
        batch_fill_map: Dict[str, Any] = {}
        next_fill_index = 0
        no_fill = PatternFill(fill_type=None)
        new_entry_fill = PatternFill(
            fill_type="solid", start_color=NEW_ENTRY_FILL_COLOR, end_color=NEW_ENTRY_FILL_COLOR
        )
        for row_idx in range(2, sheet.max_row + 1):
            rendered_value = str(sheet.cell(row=row_idx, column=rendered_at_col).value or "").strip()
            if not rendered_value:
                fill = no_fill
            else:
                if rendered_value not in batch_fill_map:
                    color = batch_fill_palette[next_fill_index % len(batch_fill_palette)]
                    batch_fill_map[rendered_value] = PatternFill(
                        fill_type="solid",
                        start_color=color,
                        end_color=color,
                    )
                    next_fill_index += 1
                fill = batch_fill_map[rendered_value]

            for col_idx in range(1, len(headers) + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = fill

        try:
            workbook.save(report_path)
            print(f"Formatted: {report_path}  (rendered_at populated: {populated} rows, color groups: {len(batch_fill_map)})")
        except PermissionError:
            print(f"[SKIP] {report_path} — file is open in another program, close it and re-run.")

    if found == 0:
        print(f"No report files found in '{root_report_folder}'.")
    else:
        print(f"\nDone — {found} report(s) formatted.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Batch ID card generator")
    parser.add_argument("--csv", default=CSV_FILE, help="Path to the CSV input file")
    parser.add_argument(
        "--district",
        type=int,
        default=None,
        metavar="N",
        help=f"Process only this district number ({DISTRICT_MIN}-{DISTRICT_MAX})",
    )
    parser.add_argument(
        "--format-reports",
        action="store_true",
        help="Format all existing district report files (add rendered_at, recolor) then exit.",
    )
    args = parser.parse_args()
    if args.format_reports:
        format_existing_reports(REPORT_FOLDER)
    else:
        batch_generate_id_cards(csv_file=args.csv, district_filter=args.district)
